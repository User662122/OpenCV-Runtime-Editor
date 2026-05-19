"""
AutoKeyPress - Excel-Driven Keyboard Automation
================================================
Reads rows from an Excel file and performs key sequences
when the trigger key (e.g. Esc) is pressed.

REQUIRED COLUMNS (any position in Excel):
  Trigger   → key to wait for before running that row (e.g. Esc, F5, Enter)
  KeyMap    → the keystroke sequence to perform for that row

ALL OTHER COLUMNS are data columns. Add or remove freely — their names
are used directly in KeyMap via v"COLNAME". Missing columns are skipped silently.

KEYMAP SYNTAX:
  v"COLNAME"       → type the value from that column (skip if column absent)
                     Special: v"Date" auto-formats the date to DDMMYY (6 digits)
                       e.g. Excel value  01-01-2025  →  types  010125
                            Excel value   5-3-2025   →  types  050325
  t"any text"      → type the literal text inside the quotes (not a column)
                     e.g. t"Hello World"  →  types  Hello World
                          t"0229skskj"    →  types  0229skskj
  Enter            → press Enter key
  Tab              → press Tab key
  Esc              → press Escape key
  Space            → press Space key
  Up/Down/Left/Right → arrow keys
  F1..F12          → function keys
  Ctrl+c           → hold Ctrl and press c  (any combo)
  Alt+F4           → hold Alt and press F4
  Shift+Tab        → hold Shift and press Tab
  2*Enter          → repeat Enter 2 times  (N*KEY)
  +                → separator between actions

EXAMPLE KeyMap:
  Enter + v"Date" + 2*Enter + v"NARRATION" + Tab + v"AMOUNT" + Enter
  Enter + t"PREFIX" + v"AMOUNT" + Tab + v"Date" + Enter
"""

import sys
import time
import re
import threading
import openpyxl
import pyautogui

# Safety: move mouse to top-left corner to abort
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.05  # 50ms between actions for reliability


# ── Key name map ────────────────────────────────────────────────────────────
KEY_MAP = {
    "enter":     "enter",
    "tab":       "tab",
    "esc":       "escape",
    "escape":    "escape",
    "space":     "space",
    "backspace": "backspace",
    "delete":    "delete",
    "del":       "delete",
    "up":        "up",
    "down":      "down",
    "left":      "left",
    "right":     "right",
    "home":      "home",
    "end":       "end",
    "pageup":    "pageup",
    "pagedown":  "pagedown",
    "insert":    "insert",
    "f1":  "f1",  "f2":  "f2",  "f3":  "f3",  "f4":  "f4",
    "f5":  "f5",  "f6":  "f6",  "f7":  "f7",  "f8":  "f8",
    "f9":  "f9",  "f10": "f10", "f11": "f11", "f12": "f12",
    "ctrl":  "ctrl",  "alt": "alt",  "shift": "shift",
    "win":   "win",   "cmd": "command",
}

MODIFIER_MAP = {
    "ctrl":  "ctrl",
    "alt":   "alt",
    "shift": "shift",
    "win":   "win",
    "cmd":   "command",
}


# ── Date formatter ───────────────────────────────────────────────────────────
def format_date_ddmmyy(value):
    """
    Convert a date value from Excel into a 6-digit DDMMYY string.

    Handles:
      - Python date/datetime objects (from openpyxl reading real date cells)
      - Strings like '01-01-2025', '1-1-2025', '01/01/2025', '1/1/25'

    Examples:
      01-01-2025  →  010125
       5-3-2025   →  050325
      01/01/25    →  010125
    """
    import datetime

    # openpyxl may return a real date/datetime object
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%d%m%y")  # zero-padded DD MM YY

    # Otherwise treat as string
    s = str(value).strip()

    # Try splitting on common separators: - / .
    for sep in ("-", "/", "."):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                day_s, mon_s, yr_s = parts[0], parts[1], parts[2]
                day = int(day_s)
                mon = int(mon_s)
                yr  = int(yr_s)
                # Normalise year: 2025 → 25, 25 → 25
                if yr >= 100:
                    yr = yr % 100
                return f"{day:02d}{mon:02d}{yr:02d}"

    # No separator found — return as-is (already numeric string?)
    return s


# ── Script parser ────────────────────────────────────────────────────────────
def parse_script(script_str):
    """
    Parse the script string into a list of action tuples:
      ('key',   key_name)
      ('combo', [mod1, mod2, ...], key_name)
      ('type',  text_value)
      ('repeat', n, sub_action)
    """
    actions = []
    tokens = [t.strip() for t in script_str.split("+")]

    i = 0
    while i < len(tokens):
        token = tokens[i]

        # ── v"COLNAME" ───────────────────────────────────────────────────
        m_val = re.fullmatch(r'v"([^"]+)"', token)
        if m_val:
            actions.append(("type_col", m_val.group(1)))
            i += 1
            continue

        # ── t"literal text" ──────────────────────────────────────────────
        m_txt = re.fullmatch(r't"([^"]*)"', token)
        if m_txt:
            actions.append(("type_literal", m_txt.group(1)))
            i += 1
            continue

        # ── N*KEY  (repeat) ──────────────────────────────────────────────
        m_rep = re.fullmatch(r'(\d+)\*(.+)', token)
        if m_rep:
            n = int(m_rep.group(1))
            sub_token = m_rep.group(2).strip()
            sub_action = parse_single_key_token(sub_token)
            if sub_action:
                actions.append(("repeat", n, sub_action))
            i += 1
            continue

        # ── Ctrl+Alt+Key  (combo with modifiers) ────────────────────────
        # Check if this token AND next token(s) form a modifier combo
        # e.g. tokens might be split at '+' so "Ctrl+c" stays one token
        # but we also handle "Ctrl" "c" as two tokens
        parts = [p.strip() for p in token.split("+") if p.strip()]
        if len(parts) == 1:
            # Could be a standalone modifier token followed by a key
            lower = parts[0].lower()
            if lower in MODIFIER_MAP and i + 1 < len(tokens):
                # Gather consecutive modifiers
                mods = []
                j = i
                while j < len(tokens) and tokens[j].lower() in MODIFIER_MAP:
                    mods.append(MODIFIER_MAP[tokens[j].lower()])
                    j += 1
                if j < len(tokens):
                    key = resolve_key(tokens[j])
                    actions.append(("combo", mods, key))
                    i = j + 1
                    continue
                else:
                    # Lone modifier, press as key
                    actions.append(("key", MODIFIER_MAP[lower]))
                    i += 1
                    continue
            # Normal single key token
            action = parse_single_key_token(token)
            if action:
                actions.append(action)
            i += 1
            continue

        # Multiple parts in one token e.g. "Ctrl+Shift+s"
        mods = []
        key = None
        for p in parts:
            pl = p.lower()
            if pl in MODIFIER_MAP:
                mods.append(MODIFIER_MAP[pl])
            else:
                key = resolve_key(p)
        if mods and key:
            actions.append(("combo", mods, key))
        elif key:
            actions.append(("key", key))
        i += 1

    return actions


def parse_single_key_token(token):
    lower = token.strip().lower()
    if lower in KEY_MAP:
        return ("key", KEY_MAP[lower])
    # Single printable character
    if len(token.strip()) == 1:
        return ("key", token.strip())
    return None


def resolve_key(token):
    lower = token.strip().lower()
    if lower in KEY_MAP:
        return KEY_MAP[lower]
    return token.strip()


# ── Action executor ──────────────────────────────────────────────────────────
def execute_actions(actions, row_data):
    for action in actions:
        kind = action[0]

        if kind == "type_col":
            col_name = action[1]
            if col_name not in row_data:
                continue  # column absent in this row's headers — skip silently
            raw_value = row_data[col_name]

            # Special handling: "Date" column → format as DDMMYY (6 digits)
            if col_name == "Date":
                value = format_date_ddmmyy(raw_value)
            else:
                value = str(raw_value)

            if not value:
                continue
            pyautogui.typewrite(value, interval=0.03)

        elif kind == "type_literal":
            text = action[1]
            if text:
                pyautogui.typewrite(text, interval=0.03)

        elif kind == "key":
            pyautogui.press(action[1])

        elif kind == "combo":
            mods, key = action[1], action[2]
            pyautogui.hotkey(*mods, key)

        elif kind == "repeat":
            n, sub = action[1], action[2]
            for _ in range(n):
                # Execute the sub-action
                execute_actions([sub], row_data)

        time.sleep(0.05)


# ── Trigger key resolver ─────────────────────────────────────────────────────
def trigger_to_key(trigger_str):
    """Convert trigger string like 'Esc' to pyautogui key name."""
    t = trigger_str.strip().lower()
    return KEY_MAP.get(t, t)


# ── Excel loader ─────────────────────────────────────────────────────────────
def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        row_dict = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        rows.append(row_dict)

    return headers, rows


# ── Trigger wait using keyboard module ───────────────────────────────────────
try:
    import keyboard
    HAS_KEYBOARD = True
except ImportError:
    HAS_KEYBOARD = False


def wait_for_trigger_keyboard(trigger_key):
    """
    Wait until the trigger key is pressed using keyboard.on_press.
    Uses on_press (not add_hotkey) because add_hotkey fails to re-register
    F-keys reliably after the first row — on_press works for every row.
    """
    event = threading.Event()

    # Normalise the key name so 'f5' matches what keyboard reports ('f5')
    target = trigger_key.strip().lower()

    def on_key(e):
        if e.name and e.name.lower() == target:
            event.set()

    # Full reset — flush any leftover hook state from the previous row
    keyboard.unhook_all()
    time.sleep(0.1)

    hook = keyboard.on_press(on_key, suppress=False)
    print(f"  → Waiting for [{trigger_key.upper()}] key...", flush=True)
    event.wait()

    keyboard.unhook_all()  # tear down immediately so next row starts clean

    # Just sleep — do NOT call keyboard.wait() here.
    # keyboard.wait re-registers an F-key listener which blocks execution.
    # A plain sleep is enough to let the key physically release.
    time.sleep(0.5)


def wait_for_trigger_input(trigger_key):
    """Fallback: press Enter in terminal to proceed."""
    input(f"  → Press ENTER in this terminal to trigger (actual key: {trigger_key.upper()})... ")


# ── Default Excel path ───────────────────────────────────────────────────────
DEFAULT_EXCEL = r"C:\Users\ROHIT DAS\Downloads\data.xlsx"


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    import argparse, os
    parser = argparse.ArgumentParser(
        description="AutoKeyPress – Excel-driven keyboard automation"
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        default=DEFAULT_EXCEL,
        help=f"Path to your .xlsx file (default: {DEFAULT_EXCEL})"
    )
    parser.add_argument(
        "--delay", type=float, default=3.0,
        help="Seconds to wait after trigger before typing (default: 3)"
    )
    parser.add_argument(
        "--sheet", type=str, default=None,
        help="Sheet name to read (default: first sheet)"
    )
    args = parser.parse_args()

    if not os.path.isfile(args.excel_file):
        print(f"\n  ERROR: File not found:")
        print(f"    {args.excel_file}")
        print(f"\n  Either:")
        print(f"    1. Save your Excel file as:  {DEFAULT_EXCEL}")
        print(f"    2. Or run with custom path:  python auto_keypress.py \"C:\\path\\to\\file.xlsx\"")
        sys.exit(1)

    print("\n" + "="*60)
    print("  AutoKeyPress – Excel Keyboard Automation")
    print("="*60)
    print(f"  File  : {args.excel_file}")
    print(f"  Delay : {args.delay}s after trigger key")
    print(f"  ABORT : Move mouse to top-left corner of screen")
    print("="*60 + "\n")

    headers, rows = load_excel(args.excel_file)

    if "Trigger" not in headers:
        print("ERROR: No 'Trigger' column found in Excel file!")
        print(f"  Found columns: {headers}")
        sys.exit(1)

    if "KeyMap" not in headers:
        print("ERROR: No 'KeyMap' column found in Excel file!")
        print(f"  Found columns: {headers}")
        sys.exit(1)

    total = len(rows)
    print(f"  Loaded {total} row(s) from Excel\n")

    for idx, row in enumerate(rows, 1):
        trigger_raw = str(row.get("Trigger", "")).strip()
        script_str  = str(row.get("KeyMap", "")).strip()

        if not trigger_raw or not script_str:
            print(f"  [Row {idx}/{total}] Skipping – no Trigger or KeyMap defined.")
            continue

        print(f"  [Row {idx}/{total}] Ready")
        print(f"    Trigger : {trigger_raw}")
        print(f"    KeyMap  : {script_str}")

        # Preview row data (exclude control columns)
        preview = {k: v for k, v in row.items() if k not in ("Trigger", "KeyMap")}
        print(f"    Data    : {preview}")

        # Wait for trigger
        trigger_key = trigger_to_key(trigger_raw)
        if HAS_KEYBOARD:
            wait_for_trigger_keyboard(trigger_key)
        else:
            wait_for_trigger_input(trigger_key)

        # Countdown
        print(f"  ✓ Triggered! Starting in {args.delay}s – switch to your target window!")
        for remaining in range(int(args.delay), 0, -1):
            print(f"    {remaining}...", end="\r", flush=True)
            time.sleep(1)
        print("  ▶ Typing now...                    ")

        # Parse and execute
        actions = parse_script(script_str)
        execute_actions(actions, row)

        print(f"  ✓ Row {idx} done.\n")
        time.sleep(0.2)

    print("="*60)
    print("  ✅ All rows processed!")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
